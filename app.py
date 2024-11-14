import tkinter as tk
from tkinter import ttk
import ttkbootstrap as tb
# Importer le fichier db.py
import db  
from fpdf import FPDF
import openpyxl

class CRUDApp:
    def __init__(self, root):
        self.root = root
        self.root.title("CRUD")
        self.root.geometry("1024x780")
        self.style = tb.Style('cosmo')

        # Variables de l'interface
        self.name_var = tk.StringVar()
        self.firstname_var = tk.StringVar()
        self.email_var = tk.StringVar()
        self.phone_var = tk.StringVar()

        self.create_widgets()
        self.populate_data()

    def create_widgets(self):
        # Frame de formulaire
        form_frame = ttk.LabelFrame(self.root, text="Informations utilisateur")
        form_frame.pack(pady=10, padx=10, fill="x")

        ttk.Label(form_frame, text="Nom :").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(form_frame, textvariable=self.name_var).grid(row=0, column=1, padx=5, pady=5)
        
        
        ttk.Label(form_frame, text="Prénom(s) :").grid(row=1, column=0, padx=5, pady=5)
        ttk.Entry(form_frame, textvariable=self.firstname_var).grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(form_frame, text="Email :").grid(row=2, column=0, padx=5, pady=5)
        ttk.Entry(form_frame, textvariable=self.email_var).grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(form_frame, text="Téléphone :").grid(row=3, column=0, padx=5, pady=5)
        ttk.Entry(form_frame, textvariable=self.phone_var).grid(row=3, column=1, padx=5, pady=5)

        # Boutons pour le CRUD
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Ajouter", command=self.add_user).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="Mettre à jour", command=self.update_user).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="Supprimer", command=self.delete_user).grid(row=0, column=2, padx=5)

        # Boutons pour l'export
        export_frame = ttk.Frame(self.root)
        export_frame.pack(pady=10)
        ttk.Button(export_frame, text="Exporter en PDF", command=self.export_to_pdf).grid(row=0, column=0, padx=5)
        ttk.Button(export_frame, text="Exporter en Excel", command=self.export_to_excel).grid(row=0, column=1, padx=5)

        # Tableau des données
        self.tree = ttk.Treeview(self.root, columns=("id", "name", "firstname", "email", "phone"), show="headings")
        self.tree.heading("id", text="ID")
        self.tree.heading("name", text="Nom")
        self.tree.heading("firstname", text="Prénom(s)")
        self.tree.heading("email", text="Email")
        self.tree.heading("phone", text="Téléphone")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self.select_user)

    def populate_data(self):
        # Vider le tableau
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Remplir le tableau avec les données
        users = db.get_all_users()
        for user in users:
            self.tree.insert('', 'end', values=user)

    def add_user(self):
        db.add_user(self.name_var.get(), self.firstname_var.get(), self.email_var.get(), self.phone_var.get())
        self.populate_data()

    def update_user(self):
        selected_item = self.tree.focus()
        if selected_item:
            user_id = self.tree.item(selected_item)['values'][0]
            db.update_user(user_id, self.name_var.get(), self.firstname_var.get(), self.email_var.get(), self.phone_var.get())
            self.populate_data()

    def delete_user(self):
        selected_item = self.tree.focus()
        if selected_item:
            user_id = self.tree.item(selected_item)['values'][0]
            db.delete_user(user_id)
            self.populate_data()

    def select_user(self, event):
        selected_item = self.tree.focus()
        if selected_item:
            values = self.tree.item(selected_item, 'values')
            self.name_var.set(values[1])
            self.firstname_var.set(values[2])
            self.email_var.set(values[3])
            self.phone_var.set(values[4])

    def export_to_pdf(self):
        users = db.get_all_users()
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.cell(200, 10, txt="Liste des Utilisateurs", ln=True, align="C")

        # En-têtes
        pdf.cell(20, 10, "ID", 1)
        pdf.cell(60, 10, "Nom", 1)
        pdf.cell(60, 10, "Prénom(s)", 1)
        pdf.cell(60, 10, "Email", 1)
        pdf.cell(40, 10, "Téléphone", 1)
        pdf.ln()

        # Données des utilisateurs
        pdf.set_font("Arial", "", 10)
        for user in users:
            pdf.cell(20, 10, str(user[0]), 1)
            pdf.cell(60, 10, user[1], 1)
            pdf.cell(60, 10, user[2], 1)
            pdf.cell(40, 10, user[3], 1)
            pdf.cell(40, 10, user[4], 1)
            pdf.ln()

        pdf.output("utilisateurs.pdf")
        print("Exporté en PDF avec succès.")

    def export_to_excel(self):
        users = db.get_all_users()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Utilisateurs"
        
        # En-têtes
        headers = ["ID", "Nom", "Prénom(s)", "Email", "Téléphone"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
        
        # Données des utilisateurs
        for row_num, user in enumerate(users, 2):
            for col_num, value in enumerate(user, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save("utilisateurs.xlsx")
        print("Exporté en Excel avec succès.")

if __name__ == "__main__":
    root = tb.Window()
    app = CRUDApp(root)
    root.mainloop()